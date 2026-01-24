import argparse
import sqlite3
from pathlib import Path

try:
    import openpyxl
except Exception as exc:  # pragma: no cover - CLI error path
    print("Missing dependency: openpyxl. Install with: pip install openpyxl")
    raise SystemExit(1) from exc


def normalize(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return " ".join(text.split())


def find_header_row(ws, max_scan: int = 50):
    for row in ws.iter_rows(min_row=1, max_row=max_scan):
        values = [normalize(cell.value).lower() for cell in row]
        if "category" in values and "brand" in values and "product" in values:
            header_map = {name: idx for idx, name in enumerate(values) if name}
            return row[0].row, header_map
    return None, None


def resolve_shop_id(cur, shop_name: str) -> int:
    cur.execute("SELECT id FROM shops WHERE name = ?", (shop_name,))
    rows = cur.fetchall()
    if not rows:
        print(f"Shop not found: {shop_name}")
        raise SystemExit(1)
    if len(rows) > 1:
        print(f"Multiple shops found with name: {shop_name}. Please use a unique name.")
        raise SystemExit(1)
    return rows[0][0]


def load_existing_map(cur, table: str, shop_id: int):
    cur.execute(f"SELECT id, name FROM {table} WHERE shop_id = ?", (shop_id,))
    data = {}
    for row_id, name in cur.fetchall():
        key = normalize(name)
        if key:
            data[key] = row_id
    return data


def main():
    parser = argparse.ArgumentParser(description="Auto-register categories, brands, and products.")
    parser.add_argument("--shop", required=True, help="Exact shop name")
    parser.add_argument("--file", required=True, help="Path to Excel (.xlsx) file")
    parser.add_argument("--sheet", default=None, help="Worksheet name (defaults to active sheet)")
    parser.add_argument("--db", default=None, help="Path to SQLite DB (defaults to instance/ezzystore.db)")
    parser.add_argument("--reorder-level", type=int, default=3, help="Default reorder level")
    parser.add_argument("--dry-run", action="store_true", help="Run without committing changes")
    args = parser.parse_args()

    repo_root = Path(__file__).resolve().parents[1]
    db_path = Path(args.db) if args.db else repo_root / "instance" / "ezzystore.db"
    if not db_path.exists():
        print(f"Database not found: {db_path}")
        raise SystemExit(1)

    file_path = Path(args.file)
    if not file_path.exists():
        print(f"Excel file not found: {file_path}")
        raise SystemExit(1)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active

    header_row, header_map = find_header_row(ws)
    if not header_row:
        print("Could not find header row with Category, Brand, Product.")
        raise SystemExit(1)

    idx_category = header_map.get("category")
    idx_brand = header_map.get("brand")
    idx_product = header_map.get("product")

    if idx_category is None or idx_brand is None or idx_product is None:
        print("Header row is missing Category, Brand, or Product columns.")
        raise SystemExit(1)

    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.cursor()

    shop_id = resolve_shop_id(cur, args.shop)

    category_map = load_existing_map(cur, "categories", shop_id)
    brand_map = load_existing_map(cur, "brands", shop_id)
    product_map = load_existing_map(cur, "products", shop_id)

    created_categories = 0
    created_brands = 0
    created_products = 0
    skipped_rows = 0

    def get_or_create_category(name: str):
        nonlocal created_categories
        key = normalize(name)
        if not key:
            return None
        if key in category_map:
            return category_map[key]
        cur.execute("INSERT INTO categories (shop_id, name) VALUES (?, ?)", (shop_id, key))
        category_id = cur.lastrowid
        category_map[key] = category_id
        created_categories += 1
        return category_id

    def get_or_create_brand(name: str):
        nonlocal created_brands
        key = normalize(name)
        if not key:
            return None
        if key in brand_map:
            return brand_map[key]
        cur.execute("INSERT INTO brands (shop_id, name) VALUES (?, ?)", (shop_id, key))
        brand_id = cur.lastrowid
        brand_map[key] = brand_id
        created_brands += 1
        return brand_id

    def get_or_create_product(name: str, brand_id, category_id):
        nonlocal created_products
        key = normalize(name)
        if not key:
            return None
        if key in product_map:
            return product_map[key]
        cur.execute(
            """
            INSERT INTO products (shop_id, brand_id, category_id, name, price, reorder_level)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (shop_id, brand_id, category_id, key, 0, args.reorder_level),
        )
        product_id = cur.lastrowid
        product_map[key] = product_id
        created_products += 1
        return product_id

    conn.execute("BEGIN")
    try:
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if row is None:
                continue
            category = row[idx_category] if idx_category < len(row) else None
            brand = row[idx_brand] if idx_brand < len(row) else None
            product = row[idx_product] if idx_product < len(row) else None

            category_name = normalize(category)
            brand_name = normalize(brand)
            product_name = normalize(product)

            if not product_name:
                skipped_rows += 1
                continue

            category_id = get_or_create_category(category_name) if category_name else None
            brand_id = get_or_create_brand(brand_name) if brand_name else None
            get_or_create_product(product_name, brand_id, category_id)

        if args.dry_run:
            conn.rollback()
        else:
            conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    print(f"Shop ID: {shop_id}")
    print(f"Created categories: {created_categories}")
    print(f"Created brands: {created_brands}")
    print(f"Created products: {created_products}")
    print(f"Skipped rows (missing product): {skipped_rows}")
    if args.dry_run:
        print("Dry run: no changes committed.")


if __name__ == "__main__":
    main()
